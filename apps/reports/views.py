"""
Vues pour les statistiques, KPIs et exports de rapports
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Count, Avg, Sum, Q, F
from django.utils import timezone
from datetime import timedelta, date, datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

from apps.accounts.models import User, Department
from apps.attendance.models import AttendanceRecord, DailyAttendance, LeaveRequest, DisciplinaryAction


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    """
    KPIs principaux pour le tableau de bord du dirigeant.
    """
    company = request.user.company
    if not company:
        return Response({'detail': 'Aucune entreprise'}, status=400)
    
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Total employés actifs
    total_employees = User.objects.filter(
        company=company, is_active_employee=True, role='employee'
    ).count()
    
    # Présences aujourd'hui
    today_present = DailyAttendance.objects.filter(
        company=company, date=today, status__in=['present', 'late', 'half_day']
    ).count()
    
    # Retards aujourd'hui
    today_late = DailyAttendance.objects.filter(
        company=company, date=today, status='late'
    ).count()
    
    # Absents aujourd'hui (employés sans pointage)
    employees_with_attendance = DailyAttendance.objects.filter(
        company=company, date=today
    ).values_list('user_id', flat=True)
    today_absent = User.objects.filter(
        company=company, is_active_employee=True, role='employee'
    ).exclude(id__in=employees_with_attendance).count()
    
    # Taux de présence du mois
    days_in_month = (today - month_start).days + 1
    expected_attendances = total_employees * days_in_month
    actual_attendances = DailyAttendance.objects.filter(
        company=company, date__gte=month_start, date__lte=today,
        status__in=['present', 'late', 'half_day']
    ).count()
    attendance_rate = (actual_attendances / expected_attendances * 100) if expected_attendances else 0
    
    # Demandes de congés en attente
    pending_leaves = LeaveRequest.objects.filter(
        company=company, status='pending'
    ).count()
    
    # Heures travaillées cette semaine
    week_hours = DailyAttendance.objects.filter(
        company=company, date__gte=week_start, date__lte=today
    ).aggregate(total=Sum('total_hours'))['total'] or 0
    
    # Répartition par département (présents aujourd'hui)
    by_department = DailyAttendance.objects.filter(
        company=company, date=today
    ).values('user__department__name', 'user__department__color').annotate(
        count=Count('id')
    )
    
    # Tendance de présence sur 7 jours
    trend_data = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        present = DailyAttendance.objects.filter(
            company=company, date=d, status__in=['present', 'late', 'half_day']
        ).count()
        late = DailyAttendance.objects.filter(
            company=company, date=d, status='late'
        ).count()
        trend_data.append({
            'date': d.isoformat(),
            'day': d.strftime('%a'),
            'present': present,
            'late': late,
            'absent': max(0, total_employees - present),
        })
    
    # Top 5 employés en retard ce mois
    top_late = AttendanceRecord.objects.filter(
        company=company,
        check_type='check_in',
        timestamp__date__gte=month_start,
        minutes_late__gt=0
    ).values(
        'user__id', 'user__first_name', 'user__last_name', 'user__avatar'
    ).annotate(
        late_count=Count('id'),
        total_minutes=Sum('minutes_late')
    ).order_by('-late_count')[:5]
    
    return Response({
        'total_employees': total_employees,
        'today': {
            'present': today_present,
            'late': today_late,
            'absent': today_absent,
            'attendance_rate': round((today_present / total_employees * 100) if total_employees else 0, 1),
        },
        'month': {
            'attendance_rate': round(attendance_rate, 1),
        },
        'week_hours': round(week_hours, 1),
        'pending_leaves': pending_leaves,
        'by_department': list(by_department),
        'trend_7days': trend_data,
        'top_late_employees': list(top_late),
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_stats(request, user_id):
    """Statistiques détaillées d'un employé"""
    try:
        employee = User.objects.get(id=user_id, company=request.user.company)
    except User.DoesNotExist:
        return Response({'detail': 'Employé introuvable'}, status=404)
    
    # Permissions
    if request.user.role == 'employee' and request.user.id != employee.id:
        return Response({'detail': 'Non autorisé'}, status=403)
    
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    daily_records = DailyAttendance.objects.filter(
        user=employee, date__gte=month_start
    )
    
    total_days = daily_records.count()
    present_days = daily_records.filter(status__in=['present', 'late', 'half_day']).count()
    late_days = daily_records.filter(status='late').count()
    absent_days = daily_records.filter(status='absent').count()
    
    total_hours = daily_records.aggregate(total=Sum('total_hours'))['total'] or 0
    avg_hours = daily_records.aggregate(avg=Avg('total_hours'))['avg'] or 0
    total_late_minutes = daily_records.aggregate(total=Sum('minutes_late'))['total'] or 0
    
    # Actions disciplinaires
    disciplinary = DisciplinaryAction.objects.filter(user=employee).count()
    
    # Congés cette année
    year_start = date(today.year, 1, 1)
    leaves_taken = LeaveRequest.objects.filter(
        user=employee, status='approved', start_date__gte=year_start
    ).count()
    
    return Response({
        'employee': {
            'id': str(employee.id),
            'name': employee.full_name,
            'role': employee.get_role_display(),
            'department': employee.department.name if employee.department else None,
        },
        'month': {
            'present_days': present_days,
            'late_days': late_days,
            'absent_days': absent_days,
            'total_hours': round(total_hours, 1),
            'avg_hours_per_day': round(avg_hours, 1),
            'total_late_minutes': total_late_minutes,
            'attendance_rate': round((present_days / total_days * 100) if total_days else 0, 1),
        },
        'disciplinary_actions': disciplinary,
        'leaves_taken_this_year': leaves_taken,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_excel(request):
    """Export Excel des présences sur une période"""
    company = request.user.company
    if not company:
        return Response({'detail': 'Aucune entreprise'}, status=400)
    
    if request.user.role == 'employee':
        return Response({'detail': 'Non autorisé'}, status=403)
    
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    
    qs = DailyAttendance.objects.filter(company=company).select_related(
        'user', 'user__department'
    )
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    qs = qs.order_by('date', 'user__last_name')
    
    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présences"
    
    # En-tête stylisé
    headers = ['Date', 'Matricule', 'Nom complet', 'Département', 'Arrivée', 'Sortie',
               'Heures', 'Retard (min)', 'Statut']
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    for row_idx, daily in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=daily.date.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=2, value=daily.user.employee_id or '-')
        ws.cell(row=row_idx, column=3, value=daily.user.full_name)
        ws.cell(row=row_idx, column=4, value=daily.user.department.name if daily.user.department else '-')
        ws.cell(row=row_idx, column=5, value=daily.check_in_time.strftime('%H:%M') if daily.check_in_time else '-')
        ws.cell(row=row_idx, column=6, value=daily.check_out_time.strftime('%H:%M') if daily.check_out_time else '-')
        ws.cell(row=row_idx, column=7, value=daily.total_hours)
        ws.cell(row=row_idx, column=8, value=daily.minutes_late)
        ws.cell(row=row_idx, column=9, value=daily.get_status_display())
        
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
    
    # Largeurs de colonnes
    widths = [12, 12, 25, 18, 10, 10, 10, 12, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Sauvegarde dans buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"presences_{company.slug}_{date_from or 'tout'}_{date_to or 'tout'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
